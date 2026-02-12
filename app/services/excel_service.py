from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.core.config import OUTPUT_DIR


def export_excel(data, file_id):

    wb = Workbook()
    ws = wb.active

    ws.title = "Income Statement"


    # ---------- Styles ----------

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    center_align = Alignment(horizontal="center")

    highlight_font = Font(bold=True)


    # ---------- Header ----------

    headers = ["Particulars"] + data.years
    ws.append(headers)

    for col in range(1, len(headers) + 1):

        cell = ws.cell(row=1, column=col)

        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align


    # ---------- Important Keywords ----------

    important_keywords = [
        "revenue",
        "total",
        "profit",
        "ebitda",
        "income",
        "expense"
    ]


    # ---------- Data Rows ----------

    for row in data.rows:

        name = row.name
        values = row.values

        excel_row = [name]

        for y in data.years:
            excel_row.append(values.get(y, "MISSING"))

        ws.append(excel_row)


        # Highlight important rows
        lower_name = name.lower()

        if any(k in lower_name for k in important_keywords):

            r = ws.max_row

            for c in range(1, len(headers) + 1):

                ws.cell(row=r, column=c).font = highlight_font


    # ---------- Freeze Header ----------

    ws.freeze_panes = "A2"


    # ---------- Auto Column Width ----------

    for col in ws.columns:

        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:

            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 3


    # ---------- Save ----------

    path = f"{OUTPUT_DIR}/{file_id}.xlsx"

    wb.save(path)

    return path
